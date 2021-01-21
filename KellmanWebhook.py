import telebot
from telebot import apihelper
from telebot import types
from openpyxl import load_workbook
from telebot import apihelper
import cherrypy
import re
import os
import datetime
from dotenv import load_dotenv
env_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(env_path):
    load_dotenv(env_path)
API_TOKEN = os.getenv('API_TOKEN_KELLMAN')


class TelegramTestKellman:
    def __init__(self, bot_ref, chat_id, file='Kellman.xlsx', criteria_number=5):
        self.bot = bot_ref
        self.answers = {i: 0 for i in ['соперничество', 'сотрудничество', 'компромисс', 'избегание', 'приспособление']}
        self.question = self.create_questions_tuple(file)
        self.state = 'START'
        self.on_test = False
        self.story = []
        self.question_counter = 0
        self.answers_base = self.create_answers_base(file)
        self.unbreak_var = None

    def change_state(self, new_state: str) -> None:
        self.state = new_state

    def validate_editing(self, new_var):
        if new_var == self.unbreak_var:
            return 0
        else:
            self.unbreak_var = new_var
            return 1

    def increment_counter(self):
        self.question_counter += 1

    def parse_from_txt(self, file):
        pass

    def parse_from_xlsx(self, file, col_number=1):
        test = {}
        wb = load_workbook(filename=file)
        sheet = wb.active
        alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        res = []
        ranges = (sheet.max_row, sheet.max_column)
        for i in range(2, ranges[0] + 1):
            cur_place = '{0}{1}'.format(alph[col_number], i)
            try:
                if len(sheet[cur_place].value) != 0:
                    res.append(sheet[cur_place].value)
            except TypeError:
                continue
        return [[res[i], res[i + 1]] for i in range(0, len(res) - 1, 2)]

    def create_questions_tuple(self, file):
        res = re.findall(r'.\w{3,5}', file)
        file_type = res[len(res) - 1]
        return {
            file_type == '.txt': self.parse_from_txt(file),
            file_type == '.xlsx': self.parse_from_xlsx(file)
        }[True]

    def create_answers_base(self, file):
        res = re.findall(r'.\w{3,5}', file)
        file_type = res[len(res) - 1]
        return {
            file_type == '.txt': self.parse_from_txt(file),
            file_type == '.xlsx': self.parse_from_xlsx(file, 2)
        }[True]

    def add_score(self, strategy):
        self.answers[strategy] += 1

    def get_question(self):
        return self.question[self.question_counter]

    def get_answer(self):
        return self.answers_base[self.question_counter]

    def send_question(self, chat_id):
        try:
            tmp = [i.split('.') for i in self.get_question()]
            text = '<b>Вопрос {0} из 30</b>\n\n<b>{1}.</b> {2}\n\n<b>{3}.</b> {4}'.format(self.question_counter + 1,
                                                                                          *tmp[0][:-1], *tmp[1][:-1])
            markup = types.InlineKeyboardMarkup(row_width=2)
            data = ['{0}{1}_{2}'.format(j, self.question_counter, i) for j, i in zip(['А', 'Б'], self.get_answer())]
            buttons = [types.InlineKeyboardButton(text=i, callback_data=j) for i, j in zip(['А', 'Б'], data)]
            markup.add(*buttons)
            self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')
        except IndexError:
            self.finish_test(chat_id)

    def get_result(self):
        return tuple(self.answers.values())

    def finish_test(self, chat_id):
        self.question_counter = 0
        text = 'Было круто! Спасибо за прохождение опросника Томаса-Килманна!\n\n<b>Ваши результаты:\n\nСоперничество - {0}\n' \
               'Сотрудничество - {1}\nКомпромисс - {2}\nИзбегание - {3}\nПриспособление - {4}\n\nУспехов в переговорах'\
               '</b>'.format(*self.get_result())
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton('Хочу пройти тест еще раз', callback_data='START_TEST_AGAIN'))
        self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')
        self.clear_answers()

    def clear_answers(self):
        for i in tuple(self.answers.keys()):
            self.answers[i] = 0

    def start(self, chat_id):
        self.on_test = True
        text = 'Приветствую!\nГотовы пройти опросник Томаса-Килманна?\nЭтот опросник поможет понять стратегии, ' \
               'которым Вы предпочитаете следовать в конфликте'
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton('Да, я хочу пройти', callback_data='START_TEST'))
        self.bot.send_message(chat_id, text, reply_markup=markup)

    def add_to_story(self, item):
        self.story.append(item)


class SimpleUserTestKellman:
    def __init__(self, msg, bot_ref):
        self.bot = bot_ref
        self.try_times = 0
        self.id = msg.chat.id
        self.test = TelegramTestKellman(bot_ref, msg.chat.id)
        self.auth(msg)

    def auth(self, msg):
        self.bot.send_message(msg.chat.id, 'Пожалуйста, введите код доступа заглавными буквами <b>без пробелов</b> и'
                                           ' знаков препинания\n\n<b>Например:</b> <i>HW2L</i>', parse_mode='html')
        self.bot.register_next_step_handler(msg, self.get_code)

    def get_code(self, msg):
        date = datetime.datetime.now()
        date += datetime.timedelta(seconds=3600 * 3)
        key = f'SW{date.month}{date.day}{str(date.year)[2::]}'
        print(key)
        if msg.text == key:
            with open('logger.txt', 'a') as logger:
                logger.write(f'[Тестобот "Томаса Киллмана"]: Пользователь с id: {msg.chat.id} Авторизация прошла успешно!'
                             f' - {datetime.datetime.today()}\n')
            self.test.change_state('START')
            self.test.start(msg.chat.id)
        else:
            self.try_times += 1
            if self.try_times <= 5:
                self.bot.send_message(msg.chat.id, f'Ой! Вы ввели неверный код.\nУточните код у тренера и повторите '
                                                   f'попытку.\n\nОсталось попыток: <b>{6 - self.try_times}</b>',
                                      parse_mode='html')
                self.bot.register_next_step_handler(msg, self.get_code)
            else:
                self.test.change_state('BLOCKED')
                with open('logger.txt', 'a') as logger:
                    logger.write(f'[Тестобот "Томаса Киллмана"]: Пользователь с id: {msg.chat.id} Ошибка авторизации!  - '
                                 f'{datetime.datetime.today()}\n')
                self.bot.send_message(msg.chat.id, 'Вы слишком много раз ввели неверный код.\n\nДля решения опишите'
                                                   ' проблему Лине, она поможет ;)\n@lina_chandler')

    def handler(self, message):
        if self.test.on_test:
            self.bot.send_message(self.id, 'Вы уже проходите тест!')
        else:
            self.test.start(self.id)

    def callback_handler(self, call):
        if call.data == 'START_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text,
                                       reply_markup=None, parse_mode='html')
            text = 'Отлично! Теперь о вопросах:\n\nСфокусируйтесь на своем поведении в переговорах на работе и в каждой' \
                   ' из следующих 30 ситуаций выберите наиболее подходящий для Вас вариант - либо А, либо Б и нажмите' \
                   ' соответствующую кнопку ниже.\n\n<b>Пожалуйста, внимательно ' \
                   'читайте вопросы, у вас не будет возможности изменить ответ.\n\nОбычно на ответы уходит до 10 ' \
                   'минут.</b>'
            self.bot.send_message(self.id, text, parse_mode='html')
            self.test.send_question(self.id)

        elif call.data == 'START_TEST_AGAIN':
            text = 'Отлично! Теперь о вопросах:\n\nСфокусируйтесь на своем поведении в переговорах на работе и в каждой' \
                   ' из следующих 30 ситуаций выберите наиболее подходящий для Вас вариант - либо А, либо Б и нажмите' \
                   ' соответствующую кнопку ниже.\n\n<b>Пожалуйста, внимательно ' \
                   'читайте вопросы, у вас не будет возможности изменить ответ.\n\nОбычно на ответы уходит до 10 ' \
                   'минут.</b>'
            self.bot.send_message(self.id, text, parse_mode='html')
            self.test.send_question(self.id)

        elif call.data[0] in ['А', 'Б']:
            strategy = call.data.split('_')[1]
            self.test.add_score(strategy)
            self.test.add_to_story(call.data[0])
            tmp = {j[0]: j[1] for j in [i.split('.')[:-1:] for i in self.test.get_question()]}
            new_text = '<b>Вопрос {0} из 30\n\n✅{1}.</b> {2}'.format(self.test.question_counter + 1, call.data[0],
                                                                    tmp[call.data[0]][1::])
            if self.test.validate_editing(call.message.message_id):
                self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                        text=new_text, reply_markup=None, parse_mode='html')
                self.test.increment_counter()
                self.test.send_question(call.message.chat.id)
            else:
                self.bot.send_message(self.id, "Пожалуйста не надо так быстро нажимать на кнопки ответа:)")


bot = telebot.TeleBot(API_TOKEN)
users = {}


@bot.message_handler(content_types='text')
def handler(message):
    try:
        if message.text != '/start':
            bot.send_message(message.chat.id, 'Вам это не нужно! Чтоб пройти тест просто нажимайте на кнопки')
        elif users[message.chat.id].test.state == 'BLOCKED':
            bot.send_message(message.chat.id, 'Вам не нужно на данный момент вводить никакой текст. Просто нажимайт'
                                              'e кнопки! :)')
        else:
            users[message.chat.id].handler(message)
    except KeyError:
        users.update([(message.chat.id, SimpleUserTestKellman(message, bot))])


@bot.callback_query_handler(func=lambda call: True)
def smo(call):
    try:
        users[call.message.chat.id].callback_handler(call)
    except KeyError:
        users.update([(call.message.chat.id, SimpleUserTestKellman(call.message, bot))])


class WebhookServer(object):
    # index равнозначно /, т.к. отсутствию части после ip-адреса (грубо говоря)
    @cherrypy.expose
    def index(self):
        length = int(cherrypy.request.headers['content-length'])
        json_string = cherrypy.request.body.read(length).decode("utf-8")
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return ''


if __name__ == '__main__':
    cherrypy.config.update({
        'server.socket_host': '127.0.0.1',
        'server.socket_port': 7775,
        'engine.autoreload.on': False
    })
    cherrypy.quickstart(WebhookServer(), '/', {'/': {}})
