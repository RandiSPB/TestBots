import telebot
from telebot import apihelper
from telebot import types
from openpyxl import load_workbook
from telebot import apihelper
import random
import re
import os
from dotenv import load_dotenv
import cherrypy
import datetime

def parse_from_xlsx_data(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    questions = {}
    tmp = None
    num = 1
    num_tmp = 0
    ranges = (sheet.max_row, sheet.max_column)
    for i in range(2, 58):
        res = []
        for j in range(ranges[1]):
            cur_place = '{0}{1}'.format(alph[j], i)
            try:
                res.append(sheet[cur_place].value)
            except TypeError:
                continue
        if num not in questions.keys():
            questions[num] = []
            questions[num].append(res)
        else:
            questions[num].append(res)
        num_tmp += 1
        if num_tmp == 2:
            num_tmp = 0
            num += 1
    return questions


class HerzbergTest:
    def __init__(self, filename, bot_ref, texts):
        self.texts = texts
        self.bot = bot_ref
        self.state = 'START'
        self.values_buffer = None
        self.generic_buffer = None
        self.bad_action = 0
        self.cur_question = 1
        self.flag = None
        self.questions = parse_from_xlsx_data(filename)
        for i in self.questions.keys():
            print(f'{i} = {self.questions[i]}')
        self.score = {'А': 0, 'Б': 0, 'В': 0, 'Г': 0, 'Д': 0, 'Е': 0, 'Ж': 0, 'З': 0}

    def change_state(self, new_state):
        self.state = new_state

    def return_cur_text(self):
        return self.texts[self.state]

    def return_cur_question_format(self, q_data):
        return f'<b>{self.cur_question}</b>. Чаще всего...\n\n<b>A. </b>{q_data[0][0]}\n<b>B. </b>{q_data[1][0]}'

    def validate_editing(self, new_args):
        pass

    def handler(self, call):

        if call.data == 'change':
            for i in self.values_buffer:
                self.score[i[0]] -= i[1]
            e_markup = types.InlineKeyboardMarkup(row_width=1)
            e_markup.add(types.InlineKeyboardButton(text='✅Изменить', callback_data='NO_CB'))
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=e_markup, parse_mode='html')
            self.send_question(call.message.chat.id)

        elif call.data == 'next_question':
            e_markup = types.InlineKeyboardMarkup(row_width=1)
            e_markup.add(types.InlineKeyboardButton(text='✅Все верно', callback_data='NO_CB'))
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=e_markup, parse_mode='html')
            self.cur_question += 1
            self.send_question(call.message.chat.id)

        elif call.data[0] == 'Q':
            key1, key2 = call.data.split('_')[1::]
            e_markup = types.InlineKeyboardMarkup(row_width=1)
            e_markup.add(types.InlineKeyboardButton(text='✅' + key1, callback_data='NO_CB'))
            question_text = 'Присвойте баллы данному утверждению от 0 до 5'
            buttons_labels = [str(i) for i in range(0, 6)]
            tmp1 = (i for i in range(0, 6))
            tmp2 = (i for i in range(5, -1, -1))
            print(tmp2)
            buttons_callback = (f'ANS_{key1}_{i}_{key2}_{j}' for i, j in zip(tmp1, tmp2))
            markup = types.InlineKeyboardMarkup(row_width=6)
            markup.add(*(types.InlineKeyboardButton(text=i, callback_data=j) for i, j in zip(buttons_labels,
                                                                                             buttons_callback)))
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=self.generic_buffer,
                                       reply_markup=e_markup, parse_mode='html')
            self.bot.send_message(call.message.chat.id, question_text, reply_markup=markup, parse_mode='html')

        elif call.data[0:3:] == 'ANS':
            key1, key1_score, key2, key2_score = call.data.split('_')[1::]
            self.score[key1] += int(key1_score)
            self.score[key2] += int(key2_score)
            self.values_buffer = ((key1, key1_score), (key2, key2_score))
            text = 'Оставшиеся баллы присвоены второму утверждению.\n\nРаспределение баллов:\n{0} - {1}\n{2} - {3}' \
                   ''.format(key1, key1_score, key2, key2_score)
            e_markup = types.InlineKeyboardMarkup(row_width=1)
            e_markup.add(types.InlineKeyboardButton(text=f'✅{key1_score}', callback_data='NO_CB'))
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=e_markup, parse_mode='html')
            self.bot.send_message(call.message.chat.id, text, reply_markup=None, parse_mode='html')
            self.cur_question += 1
            self.send_question(call.message.chat.id)
        else:
            content = (('Не надо нажимать на эту кнопку', "CAACAgIAAxkBAAEBF5NfGuC_VII-nWv3kx3vr_FG5s85dgACWwEAAhZ8aA"
                                                          "On5hm7YVB7pBoE"),
                       ('Прошу вас, не надо на нее нажимать!', 'CAACAgIAAxkBAAEBF8pfGuaFWwzW3u0jCtlkwKZJ4AfG-wACCgEAA'
                                                               'hZ8aAMGQ9g1JWM_NxoE'),
                       ('Последнее предупреждение, не надо так делать!', 'CAACAgIAAxkBAAEBF8hfGuZm4LBVGbnwmS5mYx40Sze'
                                                                         'vyAACHAEAAhZ8aANLfie-soC56hoE'),
                       ('Просто! Не надо!', 'CAACAgIAAxkBAAEBF71fGuVP6XfAvMTP7nimcsZv_6EwOAACBwEAAhZ8aANVMD-_sgABuU4aB'
                                            'A'))
            self.bot.send_sticker(call.message.chat.id, content[self.bad_action][1])
            self.bot.send_message(call.message.chat.id, content[self.bad_action][0])
            if self.bad_action < 3:
                self.bad_action += 1

    def disclaimer(self, chat_id):
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Да, все понятно', callback_data='START_TEST'))
        self.bot.send_message(chat_id, self.return_cur_text(), reply_markup=markup, parse_mode='html')

    def start(self, chat_id):
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Да, я хочу пройти', callback_data='CALL_DISC'))
        self.bot.send_message(chat_id, self.return_cur_text(), reply_markup=markup, parse_mode='html')

    def send_question(self, chat_id):
        try:
            text = f'<b>{self.cur_question} из 28</b>\n\n'
            buttons_labels = []
            for question in self.questions[self.cur_question]:
                text += f'<b>{question[1]}.</b> {question[0]}\n'
                buttons_labels.append(question[1])
            self.generic_buffer = text
            markup = types.InlineKeyboardMarkup(row_width=2)
            buttons_callback = (f'Q_{buttons_labels[0]}_{buttons_labels[1]}',
                                f'Q_{buttons_labels[1]}_{buttons_labels[0]}')
            markup.add(*(types.InlineKeyboardButton(text=i, callback_data=j) for i, j in zip(buttons_labels,
                                                                                             buttons_callback)))
            self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')
        except KeyError:
            self.finish_test(chat_id)

    def return_result(self):
        res = 'Спасибо за прохождение теста!\n<b>Ваши результаты:</b>\n\n<b>A. </b>Финансовые мотивы - <b>{0}</b>\n' \
              '<b>Б. </b>Признание и вознаграждение - <b>{1}</b>\n<b>В. </b>Ответственность работы - <b>{2}</b>\n' \
              '<b>Г. </b>Отношение с руководством - <b>{3}</b>\n<b>Д. </b>Карьера и продвижение - <b>{4}</b>\n' \
              '<b>Е. </b>Достижения и успех - <b>{5}</b>\n<b>Ж. </b>Содержание работы - <b>{6}</b>\n' \
              '<b>З. </b>Сотрудничество в коллективе - <b>{7}</b>\n\nУспехов!'.format(*self.score.values())
        return res

    def finish_test(self, chat_id):
        self.bot.send_sticker(chat_id, "CAACAgIAAxkBAAEBF5FfGuCJrOc68PUrUVsDehIeWc6dVwACQAEAAhZ8aAPOt9pjb9"
                                       "XRXRoE")
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Хочу пройти еще раз', callback_data='RESTART_TEST'))
        self.bot.send_message(chat_id, self.return_result(), reply_markup=markup, parse_mode='html')

    def start_test(self, chat_id):
        self.send_question(chat_id)


class SimpleUserHerzbergTest:
    def __init__(self, msg, bot_ref, texts):
        self.on_test = False
        self.try_times = 0
        self.bot = bot_ref
        self.id = msg.chat.id
        self.test = HerzbergTest('Herzberg.xlsx', bot_ref, texts)
        self.tmp = ('CAACAgQAAxkBAAEBF5VfGuF1UwSkKxGqd-t1XVyEAAGnaycAAuoAA9P3ri6fKevj_Rz_pRoE',
               'CAACAgIAAxkBAAEBF7tfGuUx8tRgvxEEv7VZ_FxGB3zFWwAC_Q4AAulVBRgBBl-wMEdEvBoE',
               'CAACAgIAAxkBAAEBF7lfGuTvKXFjKIBjKLHVFtvE6VHwjAACAgADkp8eETUZdc2pKL8bGgQ',
               'CAACAgIAAxkBAAEBF7dfGuTtzxFE2Z6UlmGlHcXYz1ymcAACAQADkp8eEQpfUwLsF-b2GgQ')
        self.auth(msg)

    def auth(self, msg):
        self.bot.send_message(msg.chat.id,
                              'Пожалуйста, введите код доступа заглавными буквами <b>без пробелов</b> и'
                              ' знаков препинания\n\n<b>Например:</b> <i>HW2L</i>', parse_mode='html')
        self.bot.register_next_step_handler(msg, self.get_code)

    def handler(self, message):
        if self.on_test:
            self.bot.send_message(message.chat.id, 'Вы уже проходите тест!')
        else:
            self.on_test = True
            self.test.start(message.chat.id)

    def get_code(self, msg):
        date = datetime.datetime.now()
        date += datetime.timedelta(seconds=3600 * 3)
        key = f'SW{date.month}{date.day}{str(date.year)[2::]}'
        print(key)
        if msg.text == key:
            with open('logger.txt', 'a') as logger:
                logger.write(f'[Тестобот "Мотиваторы Герцберга"]: Пользователь с id: {msg.chat.id} Авторизация прошла успешно!'
                             f' - {datetime.datetime.today()}\n')
            self.test.change_state('START')
            self.bot.send_sticker(msg.chat.id, random.choice(self.tmp))
            self.test.start(msg.chat.id)
            self.on_test = True
        else:
            self.try_times += 1
            if self.try_times <= 5:
                self.bot.send_message(msg.chat.id, f'Ой! Вы ввели неверный код.\nУточните код у тренера и повторите '
                                                   f'попытку.\n\nОсталось попыток: <b>{6 - self.try_times}</b>',
                                      parse_mode='html')
                self.bot.register_next_step_handler(msg, self.get_code)
            else:
                self.test.change_state('BLOCKED')
                with open('logger.txt', 'a') as logger:
                    logger.write(f'[Тестобот "Мотиваторы Герцберга"]: Пользователь с id: {msg.chat.id} Ошибка авторизации!  - '
                                 f'{datetime.datetime.today()}\n')
                self.bot.send_message(msg.chat.id, 'Вы слишком много раз ввели неверный код.\n\nДля решения опишите'
                                                   ' проблему Лине, она поможет ;)\n@lina_chandler')

    def callback_handler(self, call):
        print(call.data)
        if call.data == 'CALL_DISC':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=self.test.return_cur_text(), reply_markup=None, parse_mode='html')
            self.test.change_state('DISCLAIMER')
            self.test.disclaimer(call.message.chat.id)

        elif call.data == 'START_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=self.test.return_cur_text(), reply_markup=None, parse_mode='html')
            self.test.start_test(call.message.chat.id)

        elif call.data == 'RESTART_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=None, parse_mode='html')
            self.test = HerzbergTest('Herzberg.xlsx', self.bot, self.test.texts)
            self.test.start(call.message.chat.id)
        else:
            self.test.handler(call)


env_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(env_path):
    load_dotenv(env_path)
API_TOKEN = os.getenv('API_TOKEN_HERZBERG')
bot = telebot.TeleBot(API_TOKEN)
users = {}
texts_literal = {
    'DISCLAIMER': 'Отлично!\n\n<b>Пожалуйста, внимательно прочитайте эту информацию:</b>\n\nВам будет предложено '
                  '28 пар утверждений\n\nВ каждой паре утверждений распределите 5 баллов между вариантами в '
                  'зависимости от того, насколько Вы согласны с утверждениями.\n\nДля этого достаточно выбрать '
                  'одно утверждение и выбрать кол-во баллов, а второму утверждению автоматически присвоится '
                  'остаток.\n\nВозможные варианты ответа:\n5 и 0;\n4 и 1;\n3 и 2;\n2 и 3;\n1 и 4;\n0 и 5.\n\n'
                  '<b>Например:\nА. </b>Я соблюдаю традиции - <b>2</b>\n<b>Б. </b>Я нарушаю традиции - <b>3</b>',
    'START': 'Добрый день 👋🏼\n\n<b>Готовы пройти тест "Мотиваторы по Герцбергу"?</b>'
}


@bot.message_handler(content_types='text')
def handler(message):
    if message.text == '/start':
        try:
            users[message.chat.id].handler(message)
        except KeyError:
            users[message.chat.id] = SimpleUserHerzbergTest(message, bot, texts_literal)
    elif users[message.chat.id].test.state == 'BLOCKED':
        bot.send_message(message.chat.id, 'Вам не нужно на данный момент вводить никакой текст. Просто нажимайт'
                                          'e кнопки! :)')
    else:
        bot.send_message(message.chat.id, 'Вам не нужно вводить никакой текст. Просто нажимайте кнопки! :)')


@bot.callback_query_handler(func=lambda call: True)
def smo(call):
    try:
        users[call.message.chat.id].callback_handler(call)
    except KeyError as e:
        users[call.message.chat.id] = SimpleUserHerzbergTest(call.message, bot, texts_literal)


class WebhookServer(object):
    # index равнозначно /, т.к. отсутствию части после ip-адреса (грубо говоря)
    @cherrypy.expose
    def index(self):
        length = int(cherrypy.request.headers['content-length'])
        json_string = cherrypy.request.body.read(length).decode("utf-8")
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return ''


if __name__ == '__main__':
    cherrypy.config.update({
        'server.socket_host': '127.0.0.1',
        'server.socket_port': 7777,
        'engine.autoreload.on': False
    })
    cherrypy.quickstart(WebhookServer(), '/', {'/': {}})
