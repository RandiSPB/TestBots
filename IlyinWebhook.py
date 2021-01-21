import telebot
from telebot import apihelper
from telebot import types
from openpyxl import load_workbook
from telebot import apihelper
import re
import re
import os
from dotenv import load_dotenv
import cherrypy
import datetime


def parse_from_xlsx_data(data):
    pass


class IlyinTest:
    def __init__(self, filename, bot_ref, texts):
        self.texts = texts
        self.bot = bot_ref
        self.state = 'START'
        self.flag = None
        self.questions = self.create_questions_data(filename)
        self.types_scores = {
            'гнев': 0,
            'защитная реакция': 0,
            'презрение': 0,
            'вспыльчивость': 0,
            'нетерпимость': 0,
            'мстительность': 0,
            'подозрительность': 0,
            'обидчивость': 0,
        }

    def change_state(self, new_state):
        self.state = new_state

    def return_cur_text(self):
        return self.texts[self.state]

    def validate_editing(self, new_args):
        pass

    def handler(self, call):
        if call.data != 'NO_CB':
            qtype, atype, value = call.data.split('_')
            if atype == 'direct':
                self.types_scores[qtype] += int(value)
            elif atype == 'reverse':
                self.types_scores[qtype] += 3 - int(value)
            text = '<b>{0}</b>. {1}'.format(*call.message.text.split('.'))
            labels_map = {str(i): j for i, j in zip(('0', '1', '2', '3'),
                                                    ('✅Полностью не согласны', '✅Скорее не согласны',
                                                     '✅Скорее согласны', '✅Полностью согласны'))}
            markup = types.InlineKeyboardMarkup(row_width=1)
            markup.add(types.InlineKeyboardButton(text=labels_map[value], callback_data='NO_CB'))

            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=text, reply_markup=markup, parse_mode='html')
            self.send_question(call.message.chat.id)
        else:
            self.bot.send_message(call.message.chat.id, 'Не надо нажимать на эту кнопку')


    @staticmethod
    def create_questions_data(filename):
        wb = load_workbook(filename=filename)
        sheet = wb.active
        alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        questions = []
        tmp = None
        ranges = (sheet.max_row, sheet.max_column)
        for i in range(2, ranges[0] + 1):
            res = []
            for j in range(ranges[1] - 1):
                cur_place = '{0}{1}'.format(alph[j], i)
                try:
                    res.append(sheet[cur_place].value)
                except TypeError:
                    continue
            if res[3] is None:
                res[3] = 'direct'
            elif res[2] is None:
                res[2] = res[3]
                res[3] = 'reverse'
            questions.append(tuple(res))
        return questions

    def disclaimer(self, chat_id):
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Да, все понятно', callback_data='START_TEST'))
        self.bot.send_message(chat_id, self.return_cur_text(), reply_markup=markup, parse_mode='html')

    def start(self, chat_id):
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Да, я хочу пройти', callback_data='CALL_DISC'))
        self.bot.send_message(chat_id, self.return_cur_text(), reply_markup=markup, parse_mode='html')

    def send_question(self, chat_id):
        try:
            q_data = self.questions.pop(0)
            question_text = '<b>{0}</b>. {1}'.format(*q_data[0:2])
            markup = types.InlineKeyboardMarkup(row_width=1)
            button_labels = ['Полностью согласны', 'Скорее согласны', 'Скорее не согласны',
                             'Полностью не согласны']
            buttons_callback = ('{0}_{1}_{2}'.format(*q_data[2::], i) for i in range(3, -1, -1))
            markup.add(*(types.InlineKeyboardButton(text=i, callback_data=j) for i, j in zip(button_labels,
                                                                                             buttons_callback)))
            self.bot.send_message(chat_id, question_text, reply_markup=markup, parse_mode='html')
        except IndexError:
            self.finish_test(chat_id)

    def finish_test(self, chat_id):
        tmp = []
        for key in self.types_scores.keys():
            tmp.append(f'{key[0].upper() + key[1::]} - <b>{self.types_scores[key]}</b>')
        text = 'Спасибо за прохождение теста<b>\n"Личностная агрессивность и конфликтность"!\n\nВаши результаты\n\n' \
               'Активная агрессия:</b>\n{0}\n{1}\n{2}\n{3}\n\n<b>Пассивная агрессия</b>\n{4}\n{5}\n{6}\n{7}\n\n'.format(*tmp)
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(types.InlineKeyboardButton(text='Хочу пройти еще раз', callback_data='RESTART_TEST'))
        self.bot.send_message(chat_id, text, reply_markup=markup, parse_mode='html')

    def start_test(self, chat_id):
        self.send_question(chat_id)


class SimpleUserTestTeamwork:
    def __init__(self, msg, bot_ref, texts):
        self.on_test = False
        self.bot = bot_ref
        self.try_times = 0
        self.id = msg.chat.id
        self.test = IlyinTest('Ilyin.xlsx', bot_ref, texts)
        self.auth(msg)

    def auth(self, msg):
        self.bot.send_message(msg.chat.id, 'Пожалуйста, введите код доступа заглавными буквами <b>без пробелов</b> и'
                                           ' знаков препинания\n\n<b>Например:</b> <i>HW2L</i>', parse_mode='html')
        self.bot.register_next_step_handler(msg, self.get_code)

    def get_code(self, msg):
        date = datetime.datetime.now()
        date += datetime.timedelta(seconds=3600 * 3)
        key = f'SW{date.month}{date.day}{str(date.year)[2::]}'
        print(key)
        if msg.text == key:
            with open('logger.txt', 'a') as logger:
                logger.write(f'[Тестобот "Ильин"]: Пользователь с id: {msg.chat.id} Авторизация прошла успешно!'
                             f' - {datetime.datetime.today()}\n')
            self.test.change_state('START')
            self.on_test = True
            self.test.start(msg.chat.id)
        else:
            self.try_times += 1
            if self.try_times <= 5:
                self.bot.send_message(msg.chat.id, f'Ой! Вы ввели неверный код.\nУточните код у тренера и повторите '
                                                   f'попытку.\n\nОсталось попыток: <b>{6 - self.try_times}</b>',
                                      parse_mode='html')
                self.bot.register_next_step_handler(msg, self.get_code)
            else:
                self.test.change_state('BLOCKED')
                with open('logger.txt', 'a') as logger:
                    logger.write(f'[Тестобот "Ильин"]: Пользователь с id: {msg.chat.id} Ошибка авторизации!  - '
                                 f'{datetime.datetime.today()}\n')
                self.bot.send_message(msg.chat.id, 'Вы слишком много раз ввели неверный код.\n\nДля решения опишите'
                                                   ' проблему Лине, она поможет ;)\n@lina_chandler')
    def handler(self, message):
        if self.on_test:
            self.bot.send_message(message.chat.id, 'Вы уже проходите тест!')
        else:
            self.on_test = True
            self.test.start(message.chat.id)

    def callback_handler(self, call):
        if call.data == 'CALL_DISC':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=self.test.return_cur_text(), reply_markup=None, parse_mode='html')
            self.test.change_state('DISCLAIMER')
            self.test.disclaimer(call.message.chat.id)

        elif call.data == 'START_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=self.test.return_cur_text(), reply_markup=None, parse_mode='html')
            self.test.start_test(call.message.chat.id)

        elif call.data == 'RESTART_TEST':
            self.bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                       text=call.message.text, reply_markup=None, parse_mode='html')
            self.test = IlyinTest('Ilyin.xlsx', self.bot, self.test.texts)
            self.test.start(call.message.chat.id)
        else:
            self.test.handler(call)


env_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(env_path):
    load_dotenv(env_path)
API_TOKEN = os.getenv('API_TOKEN_ILYIN')
bot = telebot.TeleBot(API_TOKEN)
users = {}
texts_literal = {
        'DISCLAIMER': 'Отлично!\n\n<b>Пожалуйста, внимательно прочитайте эту информацию:</b>\n\nТест состоит из 80 '
                      'утверждений. Оцените степень своего согласия с каждым утверждением:\n\n<i>Полностью согласны\n'
                      'Скорее согласны\nСкорее не согласны\nПолностью не согласны</i>\n\nПри ответах на вопросы будьте '
                      'искренними, этот тест никто в организации не будет знать ваших личных результатов.Заполняя тест,'
                      ' вспоминайте переговорные ситуации на рабочем месте.\n\nВсе прочитали? ;)',
        'START': 'Приветствую!\n\n<b>Готовы пройти тест "Личностная агрессивность и конфликтность"?</b>\n\nЭто тест '
                 'покажет какие виды деструктивной агрессии вам свойственны и в какой степени.'
    }


@bot.message_handler(content_types='text')
def handler(message):
    if message.text == '/start':
        try:
            users[message.chat.id].handler(message)
        except KeyError:
            users[message.chat.id] = SimpleUserTestTeamwork(message, bot, texts_literal)
    elif users[message.chat.id].test.state == 'BLOCKED':
        bot.send_message(message.chat.id, 'Вам не нужно на данный момент вводить никакой текст. Просто нажимайт'
                                          'e кнопки! :)')
    else:
        bot.send_message(message.chat.id, 'Вам не нужно вводить никакой текст. Просто нажимайте кнопки! :)')


@bot.callback_query_handler(func=lambda call: True)
def smo(call):
    try:
        users[call.message.chat.id].callback_handler(call)
    except KeyError as e:
        users[call.message.chat.id] = SimpleUserTestTeamwork(call.message, bot, texts_literal)


class WebhookServer(object):
    # index равнозначно /, т.к. отсутствию части после ip-адреса (грубо говоря)
    @cherrypy.expose
    def index(self):
        length = int(cherrypy.request.headers['content-length'])
        json_string = cherrypy.request.body.read(length).decode("utf-8")
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return ''


if __name__ == '__main__':
    cherrypy.config.update({
        'server.socket_host': '127.0.0.1',
        'server.socket_port': 7774,
        'engine.autoreload.on': False
    })
    cherrypy.quickstart(WebhookServer(), '/', {'/': {}})
